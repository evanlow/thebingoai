"""Tests for GET /api/query-results/{ref}/export (CSV / Excel download, 1 MB cap)."""
import io
import openpyxl
import backend.api.query_results as qr


def _payload(columns, rows, label="orders"):
    return {"columns": columns, "rows": rows, "label": label, "row_count": len(rows)}


def test_export_csv(authenticated_client, monkeypatch):
    monkeypatch.setattr(qr, "get_query_result",
                        lambda ref, uid: _payload(["a", "b"], [[1, 2], ["x,y", 'q"z']]))
    r = authenticated_client.get("/api/query-results/abc/export?format=csv")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert 'filename="orders.csv"' in r.headers["content-disposition"]
    body = r.text
    assert body.startswith("a,b\r\n")
    assert '"x,y"' in body and '"q""z"' in body  # csv escaping


def test_export_xlsx(authenticated_client, monkeypatch):
    monkeypatch.setattr(qr, "get_query_result",
                        lambda ref, uid: _payload(["a", "b"], [[1, 2], [3, 4]]))
    r = authenticated_client.get("/api/query-results/abc/export?format=xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert 'filename="orders.xlsx"' in r.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["a", "b"]
    assert [c.value for c in ws[2]] == [1, 2]


def test_export_caps_at_1mb(authenticated_client, monkeypatch):
    big = [["x" * 200, "y" * 200] for _ in range(5000)]  # ~2 MB of CSV
    monkeypatch.setattr(qr, "get_query_result",
                        lambda ref, uid: _payload(["a", "b"], big))
    r = authenticated_client.get("/api/query-results/abc/export?format=csv")
    assert r.status_code == 200
    n = len(r.content)
    assert 0 < n <= qr.MAX_EXPORT_BYTES  # trimmed to fit


def test_export_bad_format(authenticated_client, monkeypatch):
    monkeypatch.setattr(qr, "get_query_result", lambda ref, uid: _payload(["a"], [[1]]))
    r = authenticated_client.get("/api/query-results/abc/export?format=pdf")
    assert r.status_code == 400


def test_export_expired(authenticated_client, monkeypatch):
    monkeypatch.setattr(qr, "get_query_result", lambda ref, uid: None)
    r = authenticated_client.get("/api/query-results/missing/export?format=csv")
    assert r.status_code == 404


def test_export_label_sanitized(authenticated_client, monkeypatch):
    monkeypatch.setattr(qr, "get_query_result",
                        lambda ref, uid: _payload(["a"], [[1]], label="../etc/passwd"))
    r = authenticated_client.get("/api/query-results/abc/export?format=csv")
    assert 'filename="etcpasswd.csv"' in r.headers["content-disposition"]
